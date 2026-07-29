import json
from pathlib import Path

from openpyxl import load_workbook


INPUT = Path(
    r"C:\WWF HOWEE\outputs\merge_ap_total_20260729"
    r"\Interseccion_Areas_Protegidas_Total_Clases_CSI.xlsx"
)
OUTPUT = Path(r"C:\WWF HOWEE\Geoportal_v2\assets\capas\geo_ap_metrics_m2.js")

FIELD_MAP = {
    "name": "Área protegida",
    "apHa": "Superficie AP (ha)",
    "whiteHa": "Aguas blancas (ha)",
    "clearHa": "Aguas claras (ha)",
    "mixedHa": "Aguas mixtas (ha)",
    "blackHa": "Aguas negras (ha)",
    "seasonalHa": "Inundación estacional (ha)",
    "frequentHa": "Inundación frecuente (ha)",
    "aquaticTotalHa": "Total superficie (ha)",
    "lakesHa": "Lagos y lagunas (ha)",
    "bofedalesHa": "Bofedales (combinados) (ha)",
    "river1Km": "Longitud ríos clase 1 (km)",
    "river2Km": "Longitud ríos clase 2 (km)",
    "river3Km": "Longitud ríos clase 3 (km)",
}


def number_or_none(value):
    return float(value) if isinstance(value, (int, float)) else None


workbook = load_workbook(INPUT, read_only=True, data_only=True)
sheet = workbook["Resumen"]
headers = {cell.value: cell.column for cell in sheet[4] if cell.value}
missing = [header for header in FIELD_MAP.values() if header not in headers]
if missing:
    raise RuntimeError(f"Faltan columnas requeridas: {missing}")

rows = []
for row_number, values in enumerate(sheet.iter_rows(min_row=5, values_only=True), start=5):
    name = values[headers[FIELD_MAP["name"]] - 1]
    if not name:
        continue
    record = {"name": str(name)}
    for key, header in FIELD_MAP.items():
        if key == "name":
            continue
        record[key] = number_or_none(values[headers[header] - 1])
    rows.append(record)

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
OUTPUT.write_text(
    "window.AP_METRICS_M2 = "
    + json.dumps(rows, ensure_ascii=False, separators=(",", ":"))
    + ";\n",
    encoding="utf-8",
)
print(json.dumps({"output": str(OUTPUT), "rows": len(rows)}, ensure_ascii=False))
